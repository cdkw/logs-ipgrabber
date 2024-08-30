import re
import openpyxl
import requests
from openpyxl.styles import Font
import argparse

pattern = r'\[Server thread/INFO\]: (\w+)\[/(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}):\d+\] logged'

parser = argparse.ArgumentParser(description='Extract log information.')
parser.add_argument('--file', '-f', help='Specify the log file name')
parser.add_argument('--vpn', '-v', action='store_true', help='Enable VPN finding')
parser.add_argument('--geolocation', '-g', action='store_true', help='Enable IP geolocation')
args = parser.parse_args()

if not args.file:
    print("Error: Log file name not specified. Please provide the log file using --file or -f.")
    exit()

# vpn and geolocation
if args.vpn and args.geolocation:
    access_key = 'ACCESS_KEY_HERE!' #----------------------------------------------------------------------------------------------------
elif args.vpn or args.geolocation:
    print("Error: Both VPN and geolocation options need to be specified. Please provide both or neither.")
    exit()
else:
    access_key = None




workbook = openpyxl.Workbook()
sheet = workbook.active
sheet['A1'] = 'IP'
sheet['B1'] = 'Names'
sheet['C1'] = 'VPN'
sheet['D1'] = 'Geolocation'
sheet.column_dimensions['A'].width = 17
sheet.column_dimensions['B'].width = 40
sheet.column_dimensions['C'].width = 10
sheet.column_dimensions['D'].width = 50
bold_font = Font(bold=True)
sheet['A1'].font = bold_font
sheet['B1'].font = bold_font
sheet['C1'].font = bold_font
sheet['D1'].font = bold_font


try:
    with open(args.file, 'r', encoding='utf-8') as file:
        lines = file.readlines()
except IOError:
    print(f"Error: Unable to open the file '{args.file}'")
    exit()

row = 2 

ip_data = {}


for line in lines:
    match = re.search(pattern, line)

    if match:
        name = match.group(1)
        ip = match.group(2)

        if ip in ip_data:
            if name not in ip_data[ip]['names']:
                ip_data[ip]['names'].append(name)
        else:
            ip_data[ip] = {'names': [name], 'is_vpn': None, 'geolocation': None}


if args.geolocation or args.vpn:
    for ip, data in ip_data.items():
        response = requests.get(f'http://api.ipapi.com/{ip}?access_key={access_key}')
        if response.status_code == 200:
            result = response.json()

            if args.geolocation:
                country = result.get('country_name', '')
                city = result.get('city', '')
                data['geolocation'] = f'{country}, {city}' if country or city else 'Unknown'

            if args.vpn:
                data['is_vpn'] = result.get('security', {}).get('is_vpn', False)
        else:
            print(f"Error: Unable to retrieve data for IP {ip}. Status code: {response.status_code}")
            data['geolocation'] = 'Error' if args.geolocation else '-'
            data['is_vpn'] = 'Error' if args.vpn else '-'



for ip, data in ip_data.items():
    sheet.cell(row=row, column=1).value = ip
    sheet.cell(row=row, column=2).value = ', '.join(data['names'])
    sheet.cell(row=row, column=3).value = 'Yes' if data['is_vpn'] else 'No' if args.vpn else '-'
    sheet.cell(row=row, column=4).value = data['geolocation'] if args.geolocation else '-'
    row += 1


try:
    workbook.save('output.xlsx')
    print("Extraction completed. The data has been saved to 'output.xlsx'.")
except PermissionError:
    print("Error: Unable to save the Excel file. Please make sure the file is not open in another program.")
